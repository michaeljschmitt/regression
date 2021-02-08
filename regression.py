import openpyxl, statsmodels.api as sm, numpy as np, re, os, sys, matplotlib.pyplot as plt, matplotlib.pylab as pylab, numbers, logging, shutil, itertools
from mpl_toolkits.mplot3d import Axes3D
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG,format = '%(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)

around_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

top_border = Border(bottom=Side(style='thin'))

left_border = Border(right=Side(style='thin'))

# to apply borders to a range of cells
def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

# Directory check

if os.path.exists('./Output') == False:
    os.mkdir('./Output')

# Delete previous data files

if os.path.exists('./Output/dataCleaningResults.txt'):
    os.unlink('./Output/dataCleaningResults.txt')

if os.path.exists('./Output/stepwiseRegressionResults.txt'):
    os.unlink('./Output/stepwiseRegressionResults.txt')

if os.path.exists('./Output/regressionPlot.png'):
    os.unlink('./Output/regressionPlot.png')
    
if os.path.exists('./Output/3dRegressionPlot.png'):
    os.unlink('./Output/3dRegressionPlot.png')

if os.path.exists('./Output/stepwiseRegressionExcel.xlsx'):
    os.unlink('./Output/stepwiseRegressionModelAid.xlsx')

if os.path.exists('./Output/stepwiseRegressionExcel.xlsx'):
    os.unlink('./Output/stepwiseRegressionModelAid.txt')

if os.path.exists('./ERROR.txt'):
    os.unlink('./ERROR.txt')

# Load excel file

if os.path.exists('./data.xlsx'):
    wb = openpyxl.load_workbook('./data.xlsx',data_only = True)
elif os.path.exists('./Data.xlsx'):
    wb = openpyxl.load_workbook('./Data.xlsx',data_only = True)
else:
    wb = Workbook()
    wb.save('./Data.xlsx')      

    if os.path.exists('./Readme.txt') == False:
        ReadmeFile = open('Readme.txt','w')
        ReadmeFile.write("""Welcome to my Regression Program!
Created by Michael J. Schmitt.\n
This program was designed to calculate regressions using Excel and aid you in the multiple regression modeling process.
There are a variety of outputs depending on your data set, including:
	-Full regression results
	-Stepwise regression results
	-Regression modelling aids (in .txt and .xlsx formats)
	-Data cleaning results
	-2D regression plot (for single regessor datasets)
	-3D regression plot (for two regressor datasets)

Instructions:
Please enter your data in the 'Sheet' worksheet of the newly created 'Data' workbook.
The first column is for the dependent variable data.
All subsequent columns are for independent variable data.
The first row is for header data.
Each subsequent row is a data point for your regression.
Add an asterisk symbol to beginning of any header after the first column to ignore that column in your regression.
All outputs will be saved to the 'Output' folder.
Double click the program executible once data is entered into the worksheet to run your regression analysis.
Please remember to save workbook and close all related textfiles before running program.
Thanks and enjoy!\n
Powered by Anaconda.
		""")
        ReadmeFile.close()
    
    sys.exit()

if 'Sheet' in wb.sheetnames:
    sheet = wb['Sheet']
else:    
    ReadmeFile = open('ERROR.txt','w')
    ReadmeFile.write('Please create a sheet named "Sheet" in your workbook.')
    ReadmeFile.close()
    sys.exit()

# Extract excel data

for column in range(sheet.max_column):
    if column == 0:
        y = []
        x = []
        row = 0
        exog = []
        for col in sheet.iter_cols(max_col = (column + 1)):
            for cell in col:
                if row == 0:
                    endog = cell.value
                    row = 1
                else:
                    y.append(cell.value)
    else:
        colData = []
        row = 0
        for col in sheet.iter_cols(min_col = (column + 1), max_col = (column + 1)):
            for cell in col:
                if row == 0:
                    exog.append(cell.value)
                    row = 1
                else:
                    colData.append(cell.value)
        x.insert(column-1,colData)       

# Exit and error out if sheet is blank

if (len(y) == 0) and (len(x) == 0):
    errorFile = open("ERROR.txt","w")
    errorFile.write("Worksheet is blank: please enter your data")
    errorFile.close()
    sys.exit()

# Delete columns whose headers begin with "*"

deleteCols = []

for col in range(len(x)):
    if str(exog[col]).startswith('*'):
        deleteCols.append(col)

deleteCols.sort()
deleteCols.reverse()

for col in deleteCols:
    del x[col]
    del exog[col]    

deleteCols.reverse()

# Data cleaning (remove columns with a string cell)

stringCols = []
colCount = 0
for column in x:
    for row in column:
        if ((isinstance(row,str)) == True) and ((colCount in stringCols) == False):
           stringCols.append(colCount)
    colCount += 1

stringCols.sort()     # To delete from bottom up so col number is consistant
stringCols.reverse()

for stringCol in stringCols:
    #logging.debug('Col being deleted: ' + str(stringCol))
    del exog[stringCol]
    del x[stringCol]

stringCols.reverse()

# Readjust string column numbers being deleted taking into account '*' deleted columns

if len(stringCols)>0 and len(deleteCols)>0:   
    stringCols.reverse()
    for deleteCol in deleteCols:
        logging.debug('deleteCol: ' + str(deleteCol))
        for stringCol in stringCols:
            logging.debug('deleteCol: ' + str(deleteCol))
            logging.debug('stringCols: ' + str(stringCols))
            if deleteCol <= stringCol:
                stringCols[stringCols.index(stringCol)] = stringCol + 1
    stringCols.sort()

# Data cleaning (remove rows with empty cells)

emptyRows = []

rowCount = 0
for row in y:
    if isinstance(row,numbers.Real) == False:
        emptyRows.append(rowCount)
    rowCount += 1 
            
for column in x:
    rowCount = 0
    for row in column:
        logging.debug('Row looked at to be deleted: ' + str(rowCount))
        if ((isinstance(row,numbers.Real)) == False) and ((rowCount in emptyRows) == False):
            emptyRows.append(rowCount)
            logging.debug('Row to be deleted: ' + str(rowCount))
        rowCount += 1

emptyRows.sort()     # To delete from bottom up so row number is consistant
emptyRows.reverse()

for emptyRow in emptyRows:
    del y[emptyRow]
    logging.debug('Row being deleted: ' + str(emptyRow))
    for column in range(len(x)):
        logging.debug('Number being deleted: ' + str(x[column][emptyRow]))
        del x[column][emptyRow]

emptyRows.reverse()
    
# Exit and error out if data is invalid or incomplete

if len(y) == 0:
    errorFile = open("ERROR.txt","w")
    errorFile.write("Data is invalid or incomplete: please enter only real numbers for all data points")
    errorFile.close()
    sys.exit()

# Data cleaning (check columns for constant variables)

constantCol = []

for col in range(len(x)):
    if x[col].count(x[col][0]) == len(x[col]):
        constantCol.append(col)

constantCol.sort()
constantCol.reverse()

for col in constantCol:
    del x[col]
    del exog[col]

constantCol.reverse()

# Data cleaning (record rows with data containing large deviations from the mean)

largeDevRows = []

for column in range(len(x)):
    rowCount = 0
    for row in range(len(x[column])):
        logging.debug('Col:' + str(column) + ' Row: ' + str(row) + ' Test is: ' + str(abs(x[column][row]) > (np.mean(x[column]) + 2*np.std(x[column]))))
        if (abs(x[column][row]) > (np.mean(x[column]) + 2*np.std(x[column]))) == True and ((row in largeDevRows) == False):
            logging.debug('Value to be removed: ' + str(x[column][row]))
            largeDevRows.append(rowCount)
        rowCount += 1

if len(largeDevRows) > 0:
    largeDevRows.sort()
        
logging.debug('x is: ' + str(x))
logging.debug('Rows to be removed: ' + str(largeDevRows))

# Error out if there are no regressors after data cleaning

if len(x) == 0:
    errorFile = open("ERROR.txt","w")
    errorFile.write("Data cleaning processes deleted all inputted data. Please check data.")
    errorFile.close()
    sys.exit()

# Rearranging extraction to be read by model

x2 = []
        
for row in range(len(x[0])):
    newList = []
    for col in range(len(x)):
        newList.append(x[col][row])
    x2.insert(row,newList)
        
y2 = np.asarray(y)
x2 = np.asarray(x2)

## Single regression

x2 = sm.add_constant(x2)
model = sm.OLS(y2,x2)
results = model.fit()

params = results.params

# Text formating and OLS summary output

resultsText = str(results.summary())
endogRegex = re.compile(r'\s\s\s\s\s\s\s\s\sy')
exogRegex = re.compile(r'x(\d){1,2}(\s){8,9}')

resultsText = endogRegex.sub(endog[0:10].rjust(10),resultsText)
exogCount = 0
for value in exog:
    if exogCount < 9:
        resultsText = exogRegex.sub(value[0:11].ljust(11),resultsText,1)
    else:
        resultsText = exogRegex.sub(value[0:12].ljust(12),resultsText,1) #to fix formatting issue
    exogCount += 1                              
                                  
resultsFile = open('./Output/regressionResults.txt','w',encoding='utf-8')
resultsFile.write(str(resultsText)+'\n\n')
resultsFile.close()

# Data cleaning results output file

if len(emptyRows) or len(constantCol) or len(largeDevRows) or len(stringCols) > 0 or len(deleteCols) > 0:
    cleaningResultsFile = open('./Output/dataCleaningResults.txt','w',encoding='utf-8')
    cleaningResultsFile.write('Data Cleaning Results'.center(78)+'\n')
    cleaningResultsFile.write('='*78+'\n')

    if len(emptyRows) > 0:
        emptyRows.sort()
        cleaningResultsFile.write('**The following rows were removed due to missing or invalid data:\n')
        for emptyRow in emptyRows:
            cleaningResultsFile.write(str(emptyRow+2)+'\n')

    if len(deleteCols) > 0:
        cleaningResultsFile.write('------------------------------------------------------------------------------\n**The following columns were removed due to a leading "*" character in header:\n')
        for deleteCol in deleteCols:
            cleaningResultsFile.write(str(deleteCol+2)+'\n')

    if len(stringCols) > 0:
        cleaningResultsFile.write('------------------------------------------------------------------------------\n**The following columns were removed due to the presence of string data:\n')
        for stringCol in stringCols:
            cleaningResultsFile.write(str(stringCol+2)+'\n')

    if len(largeDevRows) > 0:
        cleaningResultsFile.write('------------------------------------------------------------------------------\n**Recommend the follow rows be eliminated from dataset as they contain data outside \nacceptable variance (\u03BC > 2\u03C3):\n')
        for largeDevRow in largeDevRows:
            cleaningResultsFile.write(str(largeDevRow+2)+'\n')

    if len(constantCol) > 0:
        cleaningResultsFile.write('------------------------------------------------------------------------------\n**The following columns were removed as they were flagged as a constant variable:\n')
        for col in constantCol:
            cleaningResultsFile.write(str(col+2)+'\n')

    cleaningResultsFile.close()
    
numOfRegress = len(x)

# 2D plot

if numOfRegress == 1:
    
    constant = params[0]   
    slope = params[1]
    maxNum = max(x[0]) + (max(x[0])+min(x[0]))*0.05
    minNum = min(x[0]) - (max(x[0])+min(x[0]))*0.05
    plt.plot(x[0],y,'ro')
    plt.ylabel(endog)          
    plt.xlabel(exog[0])
    plt.plot([minNum,maxNum],[(minNum*slope)+constant,(maxNum*slope)+constant])
    pylab.savefig('./Output/regressionPlot.png',bbox_inches='tight')
    
# 3D plot

if numOfRegress == 2:
    
    constant = params[0]   
    xslope = params[1]
    zslope = params[2]

    ys = np.asarray(y)
    xs = np.asarray(x[0])
    zs = np.asarray(x[1])
    
    X,Z = np.meshgrid(xs,zs)
    
    Y = X*xslope+Z*zslope+constant
    
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    
    ax.scatter(xs,zs,ys,color='r')
    ax.plot_wireframe(X, Z, Y,color='b')
    
    ax.set_xlabel(exog[0])
    ax.set_ylabel(exog[1])
    ax.set_zlabel(endog)
    
    pylab.savefig('./Output/3dRegressionPlot.png',bbox_inches='tight')

## Stepwise regression

# Define list of regressions to be run

listOfSwitches = list(itertools.product([False,True],repeat=numOfRegress))

totalRegs = []
for regSwitch in listOfSwitches:
    regSwitch = list(regSwitch)
    singleReg = []
    listNum = 0
    regSwitch.reverse()
    for logic in regSwitch:
        if logic:
            singleReg.append(listNum)
        listNum += 1
    totalRegs.append(singleReg)
del totalRegs[0]

#if numOfRegress > 12:
#    numOfRegress = 12
#totalRegs = regressionList[numOfRegress-1]

# Determine stepwise output width (using params, as they cause len issues)
    
width = 10
for regress in totalRegs:
    
    # Format regressor data for use in model
    x3 = []
    for row in range(len(x[0])):
        newList = []
        for col in regress:
            newList.append(x[col][row])
        x3.insert(row,newList)
    
    # Define regressor headers
    newExog = []
    for i in regress:
        newExog.append(exog[i])
        
    # Create regression
    y3 = np.asarray(y)
    x3 = np.asarray(x3)
    
    x3 = sm.add_constant(x3)
    model = sm.OLS(y3,x3)
    results = model.fit()
    
    params = results.params

    for param in params:
        if len(str(round(param,4))) > width:
            width = len(str(round(param,4)))

# Create header

header = 'R^2'.center(width) + '|' + 'Constant'.ljust(width) +'|'
for exNum in range(numOfRegress):
    header = header + exog[exNum][0:width].ljust(width) + '|'

# Create text file and add title and header data

stepwiseFile = open('./Output/stepwiseRegressionResults.txt','w')
stepwiseFile.write('Stepwise Regression Results'.center((width+1)*(numOfRegress+2))+'\n'+'='*((width+1)*(numOfRegress+2))+'\n')
stepwiseFile.write(header+'\n')

# Create text file and add title and header data (excel import)

excelImport = open('./Output/stepwiseRegressionModelAid.txt','w')
excelImport.write('Stepwise Regression Results (R^2 and p-values only)'.center((width+1)*(numOfRegress+2))+'\n'+'='*((width+1)*(numOfRegress+2))+'\n')
excelImport.write(header+'\n')

# Create excel file and add title and header data

stepwiseExcel = Workbook()
stepwiseSheet = stepwiseExcel.active
stepwiseSheet["A1"] = "Stepwise Regression Results (R^2 and p-values only)"
stepwiseSheet.merge_cells(start_row=1,start_column=1,end_row=1,end_column=numOfRegress+2)
style_range(stepwiseSheet,'A1:{}1'.format(get_column_letter(numOfRegress+2)),border=around_border,font=Font(bold=True))

stepwiseSheet["A2"] = "R^2"
stepwiseSheet["B2"] = "Constant"
for exNum in range(numOfRegress):
    stepwiseSheet.cell(column=exNum+3, row=2, value=exog[exNum][0:width])
style_range(stepwiseSheet,'A2:{}2'.format(get_column_letter(numOfRegress+2)),border=top_border)

# Run various regressions

for regress in totalRegs:
    
    # Format regressor data for use in model
    x3 = []
    for row in range(len(x[0])):
        newList = []
        for col in regress:
            newList.append(x[col][row])
        x3.insert(row,newList)
    
    # Define regressor headers
    newExog = []
    for i in regress:
        newExog.append(exog[i])
        
    # Create regression
    y3 = np.asarray(y)
    x3 = np.asarray(x3)
    
    x3 = sm.add_constant(x3)
    model = sm.OLS(y3,x3)
    results = model.fit()
    
    pvalues = results.pvalues
    params = results.params
    
    # Print results
    
    # Print line of results
    
    resultsLine1 = str(round(results.rsquared,4)).ljust(width)+'|' + str(round(params[0],4)).ljust(width)+'|'
    resultsLine2 = ' '*width +'|' + str(round(pvalues[0],4)).ljust(width)+'|'

    headerCount = numOfRegress
    count = 1
    for headerNum in range(headerCount):
        if regress.count(headerNum) == 1:
            resultsLine1 += str(round(params[count],4)).ljust(width) + '|'
            resultsLine2 += str(round(pvalues[count],4)).ljust(width) + '|'
            count += 1
        else:
            resultsLine1 += ' '*width +'|'
            resultsLine2 += ' '*width +'|'
         
    stepwiseFile.write('-'*((width+1)*(numOfRegress+2))+'\n')
    stepwiseFile.write(resultsLine1+'\n')
    stepwiseFile.write(resultsLine2+'\n')

    # Print line of results (excel import)

    resultsLine1 = str(round(results.rsquared,4)).ljust(width)+'|' + str(round(pvalues[0],4)).ljust(width)+'|'

    headerCount = numOfRegress
    count = 1
    for headerNum in range(headerCount):
        if regress.count(headerNum) == 1:
            resultsLine1 += str(round(pvalues[count],4)).ljust(width) + '|'
            count += 1
        else:
            resultsLine1 += ' '*width +'|'
         
    excelImport.write('-'*((width+1)*(numOfRegress+2))+'\n')
    excelImport.write(resultsLine1+'\n')

    # Print line of results (excel sheet)

    resultsLine = [round(results.rsquared,4),round(pvalues[0],4)]

    headerCount = numOfRegress
    count = 1
    for headerNum in range(headerCount):
        if regress.count(headerNum) == 1:
            resultsLine.append(round(pvalues[count],4))
            count += 1
        else:
            resultsLine.append("")
    stepwiseSheet.append(resultsLine)

style_range(stepwiseSheet,'A2:A{}'.format((stepwiseSheet.max_row)),border=left_border)
style_range(stepwiseSheet,'A2:{0}{1}'.format(get_column_letter(stepwiseSheet.max_column),stepwiseSheet.max_row),border=around_border)

stepwiseFile.close()
excelImport.close()
stepwiseExcel.save("./Output/stepwiseRegressionModelAid.xlsx")